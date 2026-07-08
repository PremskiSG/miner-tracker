"""Compare extracted Sotkamo quarterly metrics against the user's manual
tracking in soto.xlsx (Sotkamo sheet, rows 206-232).

Usage: .venv/bin/python scripts/validate_vs_xlsx.py [--xlsx ~/Downloads/soto.xlsx]

The xlsx stores mUSD values converted at the user's manual FX rates; we convert
extracted reported-currency values with our yfinance quarterly averages, so
monetary rows get a wider tolerance than operational rows.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from miner_tracker import db
from miner_tracker.queries import metrics_long

# xlsx column -> (our metric, xlsx unit -> our USD unit factor, rel tolerance)
CHECKS = {
    "C": ("revenue", 1e6, 0.10),                 # mUSD
    "E": ("net_income", 1e6, 0.15),
    "F": ("equity", 1e6, 0.10),
    "G": ("silver_price_realized", 1.0, 0.10),
    "H": ("silver_production_oz", 1e6, 0.02),    # xlsx in Moz
    "I": ("head_grade_gpt", 1.0, 0.05),
    "J": ("tpd_derived", 1.0, 0.10),
    "K": ("aisc_derived", 1.0, 0.15),
    "M": ("recovery_pct", 0.01, 0.05),           # xlsx fraction vs our percent
}


def xlsx_rows(path: Path) -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sotkamo"]
    out = {}
    for row in range(207, 233):
        date = ws[f"B{row}"].value
        if date is None:
            continue
        period = f"{date.year}-Q{(date.month - 1) // 3 + 1}"
        vals = {}
        for col, (metric, factor, _tol) in CHECKS.items():
            v = ws[f"{col}{row}"].value
            if isinstance(v, (int, float)):
                vals[metric] = float(v) * factor
        out[period] = vals
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(Path.home() / "Downloads/soto.xlsx"))
    ap.add_argument("--company", default="NORDIC_SOSI1")
    args = ap.parse_args()

    conn = db.connect()
    market, ticker = args.company.split("_", 1)
    row = conn.execute("SELECT id FROM companies WHERE market=? AND ticker=?",
                       (market, ticker)).fetchone()
    if not row:
        print("company not found — run sync first")
        return 1
    long = metrics_long(conn, row["id"])
    ours = {(r.period, r.metric): r.value_usd for r in long.itertuples()}
    # recovery is stored as percent (e.g. 83); xlsx fraction * 0.01 factor was
    # applied above, so bring ours to the same scale
    manual = xlsx_rows(Path(args.xlsx).expanduser())

    tol_by_metric = {m: t for _, (m, _f, t) in CHECKS.items()}
    n_ok = n_flag = n_missing = 0
    print(f"{'period':9} {'metric':24} {'xlsx':>15} {'extracted':>15} {'dev':>8}")
    for period in sorted(manual):
        for metric, ref in sorted(manual[period].items()):
            got = ours.get((period, metric))
            if metric == "recovery_pct" and got is not None:
                got = got / 100.0
            if got is None:
                # only report missing for periods we actually extracted
                if any(p == period for (p, _m) in ours):
                    print(f"{period:9} {metric:24} {ref:15,.2f} {'MISSING':>15}")
                    n_missing += 1
                continue
            dev = abs(got - ref) / abs(ref) if ref else float("inf")
            flag = "" if dev <= tol_by_metric[metric] else "  <-- FLAG"
            if flag:
                n_flag += 1
            else:
                n_ok += 1
            print(f"{period:9} {metric:24} {ref:15,.2f} {got:15,.2f} {dev:7.1%}{flag}")
    print(f"\n{n_ok} ok, {n_flag} flagged (> tolerance), {n_missing} missing")
    return 0


if __name__ == "__main__":
    sys.exit(main())
